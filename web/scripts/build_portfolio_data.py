from __future__ import annotations

import argparse
import html
import json
import re
import time
import unicodedata
import urllib.parse
import urllib.request
from datetime import UTC, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
WEB = ROOT / "web"
ASSET_ROOT = WEB / "assets" / "games"
DATA_DIR = WEB / "data"
SOURCE_WORKBOOK = ROOT / "portfolio_source.xlsx"

FEATURED_NAMES = [
    "Shopping Jam",
    "Perfect Pour 3D",
    "Merge and Spin",
    "Match Stackers",
    "Block Quest Jam",
    "Liquid Links",
]

CONTRIBUTORS = {
    "irem": {
        "name": "Irem Salk",
        "role": "Game Designer, Level Designer, Ideator",
        "focus": ["Game design", "Level design", "Ideation"],
    },
    "eralp": {
        "name": "Eralp Ozer",
        "role": "Game Developer",
        "focus": ["Gameplay development", "Systems", "Mobile implementation"],
    },
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")
    return slug or "game"


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def date_iso(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return clean_text(value)


def numeric_or_none(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = clean_text(value)
    if not text or text.upper() == "N/A":
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def detect_platform(link: str, explicit: str = "") -> str:
    if "play.google.com" in link:
        return "Google Play"
    if "apps.apple.com" in link:
        return "App Store"
    return explicit or "Store"


def extract_store_id(link: str, fallback: str = "") -> str:
    if "apps.apple.com" in link:
        match = re.search(r"/id(\d+)", link)
        if match:
            return match.group(1)
    if "play.google.com" in link:
        parsed = urllib.parse.urlparse(link)
        qs = urllib.parse.parse_qs(parsed.query)
        if qs.get("id"):
            return qs["id"][0]
    return clean_text(fallback)


def country_from_app_store_url(link: str) -> str:
    match = re.search(r"apps\.apple\.com/([a-z]{2})/", link)
    return match.group(1) if match else "us"


def request_url(url: str, timeout: int = 30) -> bytes:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read()


def load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK, data_only=True)
    rows: list[dict] = []

    sheet_specs = [
        ("Irem Salk-Product", "irem", "App Store Link", "App Store"),
        ("Eralp Ozer - Developer", "eralp", "Store Link", ""),
        ("Eralp Özer - Developer", "eralp", "Store Link", ""),
    ]

    for sheet_name, contributor_key, link_header, default_platform in sheet_specs:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [clean_text(cell.value) for cell in ws[3]]
        header_index = {name: index for index, name in enumerate(headers) if name}

        for values in ws.iter_rows(min_row=4, values_only=True):
            if not any(values):
                continue

            def get(header: str):
                index = header_index.get(header)
                return values[index] if index is not None and index < len(values) else None

            link = clean_text(get(link_header))
            name = clean_text(get("Game Name"))
            if not name or not link:
                continue

            explicit_platform = clean_text(get("Platform")) or default_platform
            platform = detect_platform(link, explicit_platform)

            seller = clean_text(get("Developer/Seller"))
            release_date = get("Release Date")
            version_count = get("Version Count")
            store_id_value = get("Store ID / Package ID") or get("App Store ID")

            if (
                link_header == "Store Link"
                and platform == "App Store"
                and explicit_platform
                and explicit_platform not in {"App Store", "Google Play"}
            ):
                seller = explicit_platform
                release_date = get("Developer/Seller")
                version_count = get("Release Date")
                store_id_value = get("Version Count") or get("Store ID / Package ID")

            store_id = extract_store_id(link, store_id_value)

            rows.append(
                {
                    "name": name,
                    "slug": slugify(name),
                    "year": numeric_or_none(get("Year")),
                    "storeLink": link,
                    "platform": platform,
                    "seller": seller,
                    "releaseDate": date_iso(release_date),
                    "versionCount": numeric_or_none(version_count),
                    "storeId": store_id,
                    "contributors": [CONTRIBUTORS[contributor_key]],
                    "sourceSheets": [sheet_name],
                }
            )

    return rows


def merge_rows(rows: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for row in rows:
        key = f"{row['platform']}:{row['storeId'] or row['storeLink']}"
        if key not in merged:
            merged[key] = row.copy()
            merged[key]["contributors"] = [row["contributors"][0].copy()]
            continue

        target = merged[key]
        existing_names = {person["name"] for person in target["contributors"]}
        for contributor in row["contributors"]:
            if contributor["name"] not in existing_names:
                target["contributors"].append(contributor.copy())

        target["sourceSheets"] = sorted(set(target["sourceSheets"] + row["sourceSheets"]))
        if row["year"] and (not target["year"] or row["year"] < target["year"]):
            target["year"] = row["year"]
        if row["versionCount"] and (
            not target["versionCount"] or row["versionCount"] > target["versionCount"]
        ):
            target["versionCount"] = row["versionCount"]

    games = list(merged.values())
    games.sort(key=lambda item: ((item["year"] or 9999), item["name"].lower()))
    return games


def fetch_app_store_metadata(game: dict) -> dict:
    country = country_from_app_store_url(game["storeLink"])
    countries = [country] + [code for code in ["us", "tr"] if code != country]
    for code in countries:
        url = f"https://itunes.apple.com/lookup?id={game['storeId']}&country={code}"
        data = json.loads(request_url(url).decode("utf-8"))
        if data.get("resultCount"):
            result = data["results"][0]
            screenshots = (result.get("screenshotUrls") or result.get("ipadScreenshotUrls") or [])[:4]
            if not screenshots:
                screenshots = fetch_app_store_web_screenshots(result.get("trackViewUrl") or game["storeLink"])
            return {
                "metadataSource": "Apple iTunes Lookup API",
                "title": result.get("trackName") or game["name"],
                "description": result.get("description") or "",
                "genres": result.get("genres") or [],
                "iconUrl": result.get("artworkUrl512") or result.get("artworkUrl100"),
                "screenshots": screenshots[:4],
                "storeUrl": result.get("trackViewUrl") or game["storeLink"],
                "rating": result.get("averageUserRating"),
                "ratingCount": result.get("userRatingCount"),
                "bundleId": result.get("bundleId"),
            }
        time.sleep(0.15)
    return {}


def fetch_app_store_web_screenshots(store_url: str) -> list[str]:
    text = request_url(store_url).decode("utf-8", errors="ignore")
    text = html.unescape(text)
    raw_urls = re.findall(
        r"https://[^\"'\\\s<>]+mzstatic\.com/image/thumb/[^\"'\\\s<>]+",
        text,
        flags=re.I,
    )

    screenshots: list[str] = []
    seen_bases: set[str] = set()
    for raw_url in raw_urls:
        image_url = raw_url.rstrip(");,")
        if any(blocked in image_url for blocked in ["Placeholder", "AppIcon", "/Features"]):
            continue
        dimension_match = re.search(r"/(\d{3,4})x(\d{3,4})bb", image_url)
        if not dimension_match:
            continue
        width, height = [int(part) for part in dimension_match.groups()]
        if width < 250 or height < 250:
            continue
        base = re.sub(r"/\d{3,4}x\d{3,4}bb(?:-\d+)?\.(webp|jpg|png)$", "", image_url)
        if base in seen_bases:
            continue
        seen_bases.add(base)
        screenshots.append(image_url)
        if len(screenshots) >= 4:
            break

    return screenshots


def fetch_google_play_metadata(game: dict) -> dict:
    parsed = urllib.parse.urlparse(game["storeLink"])
    qs = urllib.parse.parse_qs(parsed.query)
    package_id = qs.get("id", [game["storeId"]])[0]
    url = f"https://play.google.com/store/apps/details?id={urllib.parse.quote(package_id)}&hl=en&gl=US"
    text = request_url(url).decode("utf-8", errors="ignore")
    text = html.unescape(text)

    title = game["name"]
    title_match = re.search(r"<title>(.*?)</title>", text, flags=re.I | re.S)
    if title_match:
        title = re.sub(r"\s+-\s+Apps on Google Play\s*$", "", title_match.group(1)).strip()

    description = ""
    desc_match = re.search(
        r'<meta\s+name="description"\s+content="(.*?)"', text, flags=re.I | re.S
    )
    if desc_match:
        description = desc_match.group(1).strip()

    og_image = ""
    og_match = re.search(
        r'<meta\s+property="og:image"\s+content="(https://play-lh\.googleusercontent\.com/[^"]+)"',
        text,
        flags=re.I,
    )
    if og_match:
        og_image = og_match.group(1)

    raw_images = re.findall(r"https://play-lh\.googleusercontent\.com/[^\s\"'<>\\]+", text)
    images: list[str] = []
    for image_url in raw_images:
        image_url = image_url.replace("\\u003d", "=").replace("\\u0026", "&")
        if image_url not in images:
            images.append(image_url)

    og_base = re.sub(r"=.*$", "", og_image) if og_image else ""
    screenshot_by_base: dict[str, tuple[int, str]] = {}
    for image_url in images:
        if image_url == og_image or "pc0xffffff" in image_url:
            continue

        dimension_match = re.search(r"=w(\d{3,4})-h(\d{3,4})(?:-[^\"'<>\\\s]+)?$", image_url)
        if not dimension_match:
            continue

        width, height = [int(part) for part in dimension_match.groups()]
        aspect = width / height if height else 0
        if min(width, height) < 250:
            continue
        if 0.82 <= aspect <= 1.22:
            continue
        if not (0.5 <= aspect <= 1.9):
            continue

        base = re.sub(r"=w\d{3,4}-h\d{3,4}(?:-[^\"'<>\\\s]+)?$", "", image_url)
        if og_base and base == og_base:
            continue
        area = width * height
        existing = screenshot_by_base.get(base)
        if not existing or area > existing[0]:
            screenshot_by_base[base] = (area, image_url)

    screenshots = [item[1] for item in screenshot_by_base.values()][:4]

    return {
        "metadataSource": "Google Play Store page",
        "title": title or game["name"],
        "description": description,
        "genres": ["Games"],
        "iconUrl": og_image or (images[0] if images else ""),
        "screenshots": screenshots,
        "storeUrl": game["storeLink"],
        "packageId": package_id,
    }


def extension_from_response(url: str, content_type: str) -> str:
    lower_url = url.lower()
    for ext in [".png", ".jpg", ".jpeg", ".webp"]:
        if ext in lower_url:
            return ".jpg" if ext == ".jpeg" else ext
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    return ".jpg"


def download_image(url: str, folder: Path, stem: str, refresh: bool) -> str:
    if not url:
        return ""

    folder.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        content_type = response.headers.get("Content-Type", "")
        ext = extension_from_response(url, content_type)
        destination = folder / f"{stem}{ext}"
        if destination.exists() and not refresh:
            if destination.stat().st_size < 2048:
                return ""
            return destination.relative_to(WEB).as_posix()
        data = response.read()
        if len(data) < 2048:
            return ""
        destination.write_bytes(data)
        return destination.relative_to(WEB).as_posix()


def apply_metadata_and_assets(games: list[dict], refresh: bool) -> list[dict]:
    existing_by_key: dict[str, dict] = {}
    existing_path = DATA_DIR / "games.json"
    if existing_path.exists():
        try:
            existing_payload = json.loads(existing_path.read_text(encoding="utf-8"))
            for existing_game in existing_payload.get("games", []):
                key = f"{existing_game.get('platform')}:{existing_game.get('storeId') or existing_game.get('storeLink')}"
                existing_by_key[key] = existing_game
        except Exception:
            existing_by_key = {}

    for index, game in enumerate(games, start=1):
        print(f"[{index:02}/{len(games)}] {game['name']} ({game['platform']})")
        metadata = {}
        existing_key = f"{game['platform']}:{game['storeId'] or game['storeLink']}"
        existing_game = existing_by_key.get(existing_key, {})
        try:
            if game["platform"] == "App Store" and game["storeId"]:
                metadata = fetch_app_store_metadata(game)
            elif game["platform"] == "Google Play":
                metadata = fetch_google_play_metadata(game)
        except Exception as exc:
            print(f"  metadata failed: {exc}")

        game["metadataSource"] = metadata.get("metadataSource") or existing_game.get("metadataSource", "")
        game["title"] = metadata.get("title") or existing_game.get("title") or game["name"]
        game["description"] = compact_description(
            metadata.get("description") or existing_game.get("description", "")
        )
        game["genres"] = metadata.get("genres") or existing_game.get("genres") or []
        game["storeUrl"] = metadata.get("storeUrl") or existing_game.get("storeUrl") or game["storeLink"]
        game["rating"] = metadata.get("rating", existing_game.get("rating"))
        game["ratingCount"] = metadata.get("ratingCount", existing_game.get("ratingCount"))
        game["bundleId"] = metadata.get("bundleId", existing_game.get("bundleId"))
        game["packageId"] = metadata.get("packageId", existing_game.get("packageId"))
        game["featured"] = game["name"] in FEATURED_NAMES

        folder = ASSET_ROOT / game["slug"]
        game["icon"] = existing_game.get("icon", "")
        if metadata.get("iconUrl"):
            try:
                game["icon"] = download_image(metadata.get("iconUrl", ""), folder, "icon", refresh)
            except Exception as exc:
                print(f"  icon failed: {exc}")

        local_screenshots = []
        for shot_index, url in enumerate((metadata.get("screenshots") or [])[:4], start=1):
            try:
                local = download_image(url, folder, f"screenshot-{shot_index}", refresh)
                if local:
                    local_screenshots.append(local)
            except Exception as exc:
                print(f"  screenshot {shot_index} failed: {exc}")
        game["screenshots"] = local_screenshots or existing_game.get("screenshots", [])

        time.sleep(0.2)

    return games


def compact_description(description: str) -> str:
    text = re.sub(r"\s+", " ", description or "").strip()
    text = re.sub(r"Terms of Use:.*$", "", text, flags=re.I).strip()
    if len(text) <= 220:
        return text
    return text[:217].rsplit(" ", 1)[0] + "..."


def build_stats(games: list[dict]) -> dict:
    app_store = sum(1 for game in games if game["platform"] == "App Store")
    google_play = sum(1 for game in games if game["platform"] == "Google Play")
    years = sorted({game["year"] for game in games if game["year"]})
    contributors = sorted({person["name"] for game in games for person in game["contributors"]})
    total_versions = sum(game["versionCount"] or 0 for game in games)
    return {
        "totalGames": len(games),
        "featuredGames": sum(1 for game in games if game["featured"]),
        "appStoreGames": app_store,
        "googlePlayGames": google_play,
        "yearRange": f"{years[0]}-{years[-1]}" if years else "",
        "contributors": contributors,
        "totalVersionCount": total_versions,
    }


def write_data(games: list[dict]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedAt": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "sourceWorkbook": SOURCE_WORKBOOK.name,
        "contributors": list(CONTRIBUTORS.values()),
        "featuredOrder": FEATURED_NAMES,
        "stats": build_stats(games),
        "games": games,
    }
    data = json.dumps(payload, ensure_ascii=False, indent=2)
    (DATA_DIR / "games.json").write_text(data + "\n", encoding="utf-8")
    (DATA_DIR / "games.js").write_text(
        "window.PORTFOLIO_DATA = " + data + ";\n", encoding="utf-8"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true", help="Redownload existing assets")
    args = parser.parse_args()

    rows = load_rows()
    games = merge_rows(rows)
    games = apply_metadata_and_assets(games, refresh=args.refresh)
    write_data(games)
    print(f"Wrote {len(games)} games to {DATA_DIR / 'games.json'}")


if __name__ == "__main__":
    main()
